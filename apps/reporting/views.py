from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Count, Q
import csv
import io


def get_report_queryset(user):
    """Retourne le queryset pour les rapports selon le rôle."""
    from apps.tickets.models import Ticket
    from apps.accounts.models import User
    from apps.tickets.views import get_tickets_for_user
    return get_tickets_for_user(user)


@login_required
def report_index(request):
    return render(request, 'reporting/index.html')


@login_required
def report_tickets(request):
    from apps.tickets.models import Ticket

    user = request.user
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category = request.GET.get('category')
    status = request.GET.get('status')

    qs = get_report_queryset(user)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if category:
        qs = qs.filter(category__type=category)
    if status:
        qs = qs.filter(status=status)

    by_category = list(qs.values('category__name').annotate(count=Count('id')).order_by('-count'))
    by_priority = list(qs.values('priority').annotate(count=Count('id')))
    by_status = list(qs.values('status').annotate(count=Count('id')))
    by_app = list(qs.values('application__name').annotate(count=Count('id')).order_by('-count')[:10])
    by_agent = list(
        qs.filter(assigned_to__isnull=False)
        .values('assigned_to__username', 'assigned_to__first_name', 'assigned_to__last_name')
        .annotate(count=Count('id')).order_by('-count')[:10]
    )
    by_dept = list(
        qs.values('department__name').annotate(count=Count('id')).order_by('-count')[:10]
    )
    sla_respected = qs.filter(
        sla_response_exceeded=False, sla_resolution_exceeded=False,
        status__in=[Ticket.STATUS_CLOTURE, Ticket.STATUS_RESOLU]
    ).count()
    sla_exceeded = qs.filter(
        Q(sla_response_exceeded=True) | Q(sla_resolution_exceeded=True)
    ).count()
    out_of_sla = qs.filter(resolved_out_of_sla=True).count()

    context = {
        'tickets_count': qs.count(),
        'by_category': by_category,
        'by_priority': by_priority,
        'by_status': by_status,
        'by_app': by_app,
        'by_agent': by_agent,
        'by_dept': by_dept,
        'sla_stats': {
            'respected': sla_respected,
            'exceeded': sla_exceeded,
            'out_of_sla': out_of_sla,
        },
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reporting/tickets_report.html', context)


@login_required
def export_csv(request):
    from apps.tickets.models import Ticket

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="tickets_export.csv"'
    response.write('﻿')  # BOM UTF-8 pour Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Numéro', 'Titre', 'Catégorie', 'Priorité', 'Statut',
        'Demandeur', 'Département demandeur', 'Responsable',
        'Application', 'Date création', 'Date résolution',
        'SLA Prise en charge dépassée', 'SLA Traitement dépassé', 'Résolu hors SLA',
        'Temps traitement (h)'
    ])

    qs = get_report_queryset(request.user).select_related(
        'category', 'application', 'department', 'created_by', 'assigned_to'
    ).order_by('-created_at')

    for t in qs:
        writer.writerow([
            t.number, t.title,
            t.category.name if t.category else '',
            t.get_priority_display(), t.get_status_display(),
            t.created_by.get_full_name() or t.created_by.username if t.created_by else '',
            t.created_by.department.name if t.created_by and t.created_by.department else '',
            t.assigned_to.get_full_name() or t.assigned_to.username if t.assigned_to else '',
            t.application.name if t.application else '',
            t.created_at.strftime('%d/%m/%Y %H:%M') if t.created_at else '',
            t.resolved_at.strftime('%d/%m/%Y %H:%M') if t.resolved_at else '',
            'Oui' if t.sla_response_exceeded else 'Non',
            'Oui' if t.sla_resolution_exceeded else 'Non',
            'Oui' if t.resolved_out_of_sla else 'Non',
            t.processing_time,
        ])
    return response


@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = [
        'Numéro', 'Titre', 'Catégorie', 'Priorité', 'Statut',
        'Demandeur', 'Département', 'Responsable', 'Application',
        'Date création', 'Date résolution',
        'SLA PC dépassée', 'SLA Traitement dépassé', 'Résolu hors SLA', 'Temps (h)'
    ]

    hdr_fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    priority_colors = {
        'P0': 'FFCCCC', 'P1': 'FFEECC',
        'P2': 'FFFFCC', 'P3': 'CCFFCC'
    }

    qs = get_report_queryset(request.user).select_related(
        'category', 'application', 'department', 'created_by', 'assigned_to'
    ).order_by('-created_at')

    for row_num, t in enumerate(qs, 2):
        row_data = [
            t.number, t.title,
            t.category.name if t.category else '',
            t.get_priority_display(), t.get_status_display(),
            t.created_by.get_full_name() if t.created_by else '',
            t.created_by.department.name if t.created_by and t.created_by.department else '',
            t.assigned_to.get_full_name() if t.assigned_to else '',
            t.application.name if t.application else '',
            t.created_at.strftime('%d/%m/%Y %H:%M') if t.created_at else '',
            t.resolved_at.strftime('%d/%m/%Y %H:%M') if t.resolved_at else '',
            'Oui' if t.sla_response_exceeded else 'Non',
            'Oui' if t.sla_resolution_exceeded else 'Non',
            'Oui' if t.resolved_out_of_sla else 'Non',
            t.processing_time,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            if col == 4:  # colonne Priorité
                color = priority_colors.get(t.priority, 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            if t.resolved_out_of_sla and col == 14:
                cell.font = Font(color='FF0000', bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tickets_export.xlsx"'
    return response
