"""
Génère le fichier Excel de test pour l'import des employés autorisés.
Usage : python generate_test_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employés autorisés"

# En-têtes
headers = ["CUID", "Statut", "Département"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")  # Orange

for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Données de test
rows = [
    ("VFYX5401", "permanent", "Informatique"),
    # Exemples supplémentaires (à titre indicatif)
    # ("AB12345", "stagiaire", "Comptabilité"),
    # ("CD67890", "interimaire", "Ressources Humaines"),
]

for row_idx, (cuid, statut, dept) in enumerate(rows, start=2):
    ws.cell(row=row_idx, column=1, value=cuid)
    ws.cell(row=row_idx, column=2, value=statut)
    ws.cell(row=row_idx, column=3, value=dept)

# Largeurs de colonnes
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 25

output = Path("employes_autorises_test.xlsx")
wb.save(output)
print(f"Fichier généré : {output.resolve()}")
